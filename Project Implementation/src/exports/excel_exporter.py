"""
ExcelExporter - Exports data to Excel format using openpyxl.
"""

from datetime import datetime
from typing import Dict, List, Optional, Any
from pathlib import Path
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, LineChart, Reference

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import APP_NAME, CURRENCY_SYMBOL


class ExcelExporter:
    """
    Exports expense data to Excel format.
    Supports styled tables, charts, and multiple worksheets.
    """

    def __init__(self):
        """Initialize ExcelExporter."""
        self._setup_styles()

    def _setup_styles(self):
        """Setup reusable cell styles."""
        # Header style
        self.header_font = Font(bold=True, color='FFFFFF', size=11)
        self.header_fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')

        # Data cell style
        self.data_font = Font(size=10)
        self.data_alignment = Alignment(horizontal='left', vertical='center')

        # Currency style
        self.currency_alignment = Alignment(horizontal='right', vertical='center')

        # Border style
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Alternate row fill
        self.alt_row_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

        # Title style
        self.title_font = Font(bold=True, size=14, color='2C3E50')

        # Status colors
        self.status_fills = {
            'ok': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),
            'warning': PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),
            'critical': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
            'over': PatternFill(start_color='F5C6CB', end_color='F5C6CB', fill_type='solid')
        }

    def _apply_header_style(self, cell):
        """Apply header style to cell."""
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.header_alignment
        cell.border = self.thin_border

    def _apply_data_style(self, cell, row_idx: int, is_currency: bool = False):
        """Apply data style to cell."""
        cell.font = self.data_font
        cell.alignment = self.currency_alignment if is_currency else self.data_alignment
        cell.border = self.thin_border
        if row_idx % 2 == 0:
            cell.fill = self.alt_row_fill

    def _auto_column_width(self, ws, min_width: int = 10, max_width: int = 50):
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    # ===== EXPENSE LIST EXPORT =====

    def export_expenses(
        self,
        expenses: List[Dict],
        filepath: str,
        sheet_name: str = "Expenses"
    ) -> bool:
        """
        Export list of expenses to Excel.

        Args:
            expenses: List of expense dictionaries
            filepath: Output file path
            sheet_name: Worksheet name

        Returns:
            True on success
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Title
            ws['A1'] = f"{APP_NAME} - Expense Export"
            ws['A1'].font = self.title_font
            ws.merge_cells('A1:F1')

            # Date generated
            ws['A2'] = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A2'].font = Font(italic=True, size=9)

            # Headers
            headers = ['Date', 'Vendor', 'Category', 'Subcategory', 'Description', 'Amount', 'Payment Method']
            header_row = 4

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                self._apply_header_style(cell)

            # Data
            for row_idx, expense in enumerate(expenses, 1):
                data_row = header_row + row_idx

                # Date
                date_val = expense.get('date', '')
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime('%d/%m/%Y')
                else:
                    date_str = str(date_val)

                cells_data = [
                    date_str,
                    expense.get('vendor', ''),
                    expense.get('category', ''),
                    expense.get('subcategory', ''),
                    expense.get('description', ''),
                    expense.get('amount', 0),
                    expense.get('payment_method', '')
                ]

                for col, value in enumerate(cells_data, 1):
                    cell = ws.cell(row=data_row, column=col, value=value)
                    is_currency = col == 6  # Amount column
                    self._apply_data_style(cell, row_idx, is_currency)

                    if is_currency:
                        cell.number_format = f'#,##0.00 "{CURRENCY_SYMBOL}"'

            # Total row
            if expenses:
                total_row = header_row + len(expenses) + 1
                ws.cell(row=total_row, column=5, value="Total:").font = Font(bold=True)
                total_cell = ws.cell(row=total_row, column=6,
                                    value=sum(e.get('amount', 0) for e in expenses))
                total_cell.font = Font(bold=True)
                total_cell.number_format = f'#,##0.00 "{CURRENCY_SYMBOL}"'

            self._auto_column_width(ws)
            wb.save(filepath)
            return True

        except Exception as e:
            print(f"Excel export error: {e}")
            return False

    # ===== MONTHLY REPORT EXPORT =====

    def export_monthly_report(
        self,
        report_data: Dict,
        filepath: str
    ) -> bool:
        """
        Export monthly report to Excel with multiple sheets.

        Args:
            report_data: Monthly report data dictionary
            filepath: Output file path

        Returns:
            True on success
        """
        try:
            wb = Workbook()

            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            self._create_summary_sheet(ws_summary, report_data)

            # Category breakdown sheet
            ws_categories = wb.create_sheet("By Category")
            self._create_category_sheet(ws_categories, report_data.get('by_category', []))

            # Vendor sheet
            ws_vendors = wb.create_sheet("By Vendor")
            self._create_vendor_sheet(ws_vendors, report_data.get('by_vendor', []))

            # Daily trend sheet
            ws_daily = wb.create_sheet("Daily Trend")
            self._create_daily_trend_sheet(ws_daily, report_data.get('by_day', []))

            # Budget status sheet
            if report_data.get('budget_status'):
                ws_budget = wb.create_sheet("Budget Status")
                self._create_budget_sheet(ws_budget, report_data.get('budget_status', []))

            wb.save(filepath)
            return True

        except Exception as e:
            print(f"Excel export error: {e}")
            return False

    def _create_summary_sheet(self, ws, report_data: Dict):
        """Create summary worksheet."""
        period = report_data.get('period', {})
        summary = report_data.get('summary', {})

        ws['A1'] = f"Monthly Report - {period.get('month', 'Unknown')}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:C1')

        ws['A3'] = "Key Metrics"
        ws['A3'].font = Font(bold=True, size=12)

        metrics = [
            ('Total Expenses', summary.get('total_amount', 0)),
            ('Number of Transactions', summary.get('expense_count', 0)),
            ('Average Expense', summary.get('average_amount', 0)),
            ('Highest Expense', summary.get('max_amount', 0)),
            ('Lowest Expense', summary.get('min_amount', 0)),
            ('Median Expense', summary.get('median_amount', 0))
        ]

        for row, (label, value) in enumerate(metrics, 5):
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=value)
            if isinstance(value, (int, float)) and label != 'Number of Transactions':
                cell.number_format = f'#,##0.00 "{CURRENCY_SYMBOL}"'

        self._auto_column_width(ws)

    def _create_category_sheet(self, ws, categories: List[Dict]):
        """Create category breakdown worksheet."""
        ws['A1'] = "Expenses by Category"
        ws['A1'].font = self.title_font

        headers = ['Category', 'Amount', 'Count', 'Percentage', 'Average']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_header_style(cell)

        for row_idx, cat in enumerate(categories, 1):
            row = 3 + row_idx
            ws.cell(row=row, column=1, value=cat.get('category', ''))
            ws.cell(row=row, column=2, value=cat.get('amount', 0)).number_format = f'#,##0.00 "{CURRENCY_SYMBOL}"'
            ws.cell(row=row, column=3, value=cat.get('count', 0))
            ws.cell(row=row, column=4, value=cat.get('percentage', 0) / 100).number_format = '0.0%'
            ws.cell(row=row, column=5, value=cat.get('average', 0)).number_format = f'#,##0.00 "{CURRENCY_SYMBOL}"'

        # Add pie chart
        if categories:
            chart = PieChart()
            chart.title = "Expense Distribution"
            data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(categories))
            labels = Reference(ws, min_col=1, min_row=4, max_row=3 + len(categories))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.width = 12
            chart.height = 8
            ws.add_chart(chart, "G3")

        self._auto_column_width(ws)

    def _create_vendor_sheet(self, ws, vendors: List[Dict]):
        """Create vendor breakdown worksheet."""
        ws['A1'] = "Top Vendors"
        ws['A1'].font = self.title_font

        headers = ['Vendor', 'Amount', 'Transactions', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_header_style(cell)

        for row_idx, vendor in enumerate(vendors[:20], 1):
            row = 3 + row_idx
            ws.cell(row=row, column=1, value=vendor.get('vendor', ''))
            ws.cell(row=row, column=2, value=vendor.get('amount', 0)).number_format = f'#,##0.00 "{CURRENCY_SYMBOL}"'
            ws.cell(row=row, column=3, value=vendor.get('count', 0))
            ws.cell(row=row, column=4, value=vendor.get('percentage', 0) / 100).number_format = '0.0%'

        self._auto_column_width(ws)

    def _create_daily_trend_sheet(self, ws, daily_data: List[Dict]):
        """Create daily trend worksheet."""
        ws['A1'] = "Daily Expense Trend"
        ws['A1'].font = self.title_font

        headers = ['Date', 'Day', 'Amount', 'Transactions']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_header_style(cell)

        for row_idx, day in enumerate(daily_data, 1):
            row = 3 + row_idx
            ws.cell(row=row, column=1, value=day.get('date', ''))
            ws.cell(row=row, column=2, value=day.get('day_name', ''))
            ws.cell(row=row, column=3, value=day.get('amount', 0)).number_format = f'#,##0.00 "{CURRENCY_SYMBOL}"'
            ws.cell(row=row, column=4, value=day.get('count', 0))

        # Add line chart
        if daily_data:
            chart = LineChart()
            chart.title = "Daily Trend"
            chart.y_axis.title = "Amount"
            data = Reference(ws, min_col=3, min_row=3, max_row=3 + len(daily_data))
            chart.add_data(data, titles_from_data=True)
            chart.width = 15
            chart.height = 8
            ws.add_chart(chart, "F3")

        self._auto_column_width(ws)

    def _create_budget_sheet(self, ws, budget_data: List[Dict]):
        """Create budget status worksheet."""
        ws['A1'] = "Budget Status"
        ws['A1'].font = self.title_font

        headers = ['Category', 'Budget', 'Spent', 'Remaining', 'Used %', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_header_style(cell)

        for row_idx, budget in enumerate(budget_data, 1):
            row = 3 + row_idx
            budget_amt = budget.get('budget', budget.get('amount', 0))
            spent = budget.get('actual', budget.get('spent', 0))
            remaining = budget_amt - spent
            pct = budget.get('used_percentage', (spent / budget_amt * 100) if budget_amt > 0 else 0)

            if pct >= 100:
                status = 'Over Budget'
                fill = self.status_fills['over']
            elif pct >= 80:
                status = 'Warning'
                fill = self.status_fills['warning']
            else:
                status = 'OK'
                fill = self.status_fills['ok']

            ws.cell(row=row, column=1, value=budget.get('category', ''))
            ws.cell(row=row, column=2, value=budget_amt).number_format = f'#,##0.00 "{CURRENCY_SYMBOL}"'
            ws.cell(row=row, column=3, value=spent).number_format = f'#,##0.00 "{CURRENCY_SYMBOL}"'
            ws.cell(row=row, column=4, value=remaining).number_format = f'#,##0.00 "{CURRENCY_SYMBOL}"'
            ws.cell(row=row, column=5, value=pct / 100).number_format = '0.0%'
            status_cell = ws.cell(row=row, column=6, value=status)
            status_cell.fill = fill

        # Add bar chart
        if budget_data:
            chart = BarChart()
            chart.title = "Budget vs Spent"
            chart.type = "col"
            data = Reference(ws, min_col=2, max_col=3, min_row=3, max_row=3 + len(budget_data))
            cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(budget_data))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 15
            chart.height = 8
            ws.add_chart(chart, "H3")

        self._auto_column_width(ws)

    # ===== ANNUAL REPORT EXPORT =====

    def export_annual_report(
        self,
        report_data: Dict,
        filepath: str
    ) -> bool:
        """
        Export annual report to Excel.

        Args:
            report_data: Annual report data dictionary
            filepath: Output file path

        Returns:
            True on success
        """
        try:
            wb = Workbook()

            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Annual Summary"

            period = report_data.get('period', {})
            ws_summary['A1'] = f"Annual Report - {period.get('year', 'Unknown')}"
            ws_summary['A1'].font = self.title_font

            summary = report_data.get('summary', {})
            metrics = [
                ('Total Expenses', summary.get('total_amount', 0)),
                ('Number of Transactions', summary.get('expense_count', 0)),
                ('Average Expense', summary.get('average_amount', 0)),
                ('Estimated Monthly Recurring', report_data.get('estimated_monthly_recurring', 0))
            ]

            for row, (label, value) in enumerate(metrics, 3):
                ws_summary.cell(row=row, column=1, value=label)
                cell = ws_summary.cell(row=row, column=2, value=value)
                if isinstance(value, (int, float)) and label != 'Number of Transactions':
                    cell.number_format = f'#,##0.00 "{CURRENCY_SYMBOL}"'

            # Monthly trend sheet
            ws_monthly = wb.create_sheet("Monthly Trend")
            monthly = report_data.get('monthly_trend', [])

            headers = ['Month', 'Amount', 'Transactions']
            for col, header in enumerate(headers, 1):
                cell = ws_monthly.cell(row=1, column=col, value=header)
                self._apply_header_style(cell)

            for row_idx, m in enumerate(monthly, 1):
                row = 1 + row_idx
                ws_monthly.cell(row=row, column=1, value=m.get('month_name', m.get('month', '')))
                ws_monthly.cell(row=row, column=2, value=m.get('amount', 0)).number_format = f'#,##0.00 "{CURRENCY_SYMBOL}"'
                ws_monthly.cell(row=row, column=3, value=m.get('count', 0))

            self._auto_column_width(ws_monthly)

            # Category sheet
            ws_categories = wb.create_sheet("By Category")
            self._create_category_sheet(ws_categories, report_data.get('by_category', []))

            # Vendor sheet
            ws_vendors = wb.create_sheet("By Vendor")
            self._create_vendor_sheet(ws_vendors, report_data.get('by_vendor', []))

            self._auto_column_width(ws_summary)
            wb.save(filepath)
            return True

        except Exception as e:
            print(f"Excel export error: {e}")
            return False

    def export_to_bytes(
        self,
        expenses: List[Dict]
    ) -> bytes:
        """
        Export to Excel bytes (for streaming/download).

        Args:
            expenses: List of expense dictionaries

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"

        # Headers
        headers = ['Date', 'Vendor', 'Category', 'Subcategory', 'Amount', 'Payment Method']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Data
        for row_idx, expense in enumerate(expenses, 2):
            date_val = expense.get('date', '')
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%d/%m/%Y')
            else:
                date_str = str(date_val)

            ws.cell(row=row_idx, column=1, value=date_str)
            ws.cell(row=row_idx, column=2, value=expense.get('vendor', ''))
            ws.cell(row=row_idx, column=3, value=expense.get('category', ''))
            ws.cell(row=row_idx, column=4, value=expense.get('subcategory', ''))
            ws.cell(row=row_idx, column=5, value=expense.get('amount', 0)).number_format = f'#,##0.00 "{CURRENCY_SYMBOL}"'
            ws.cell(row=row_idx, column=6, value=expense.get('payment_method', ''))

        self._auto_column_width(ws)

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


# Singleton instance
_excel_exporter: Optional[ExcelExporter] = None


def get_excel_exporter() -> ExcelExporter:
    """Get or create global ExcelExporter instance."""
    global _excel_exporter
    if _excel_exporter is None:
        _excel_exporter = ExcelExporter()
    return _excel_exporter
